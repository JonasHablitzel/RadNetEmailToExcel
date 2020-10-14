import os
import sys, getopt
import yaml
import imaplib
import email
from email.header import decode_header
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill
#import csv


# maybe csv exporter

def parse_list_toxlsx(starterlist:list,path:str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Meldungen"
    if len(starterlist)>0:
        dictkeys=list(starterlist[0].keys())
    
    ws.append(dictkeys)
    for starter in starterlist:
        ws.append(list(starter.values()))

    bold_font = Font(color='00FFFFFF', bold=True)
    greay_fill=PatternFill("solid",fgColor='00333333')
    #Format Cells
    for cell in ws["1:1"]:
        cell.fill = greay_fill
        cell.font = bold_font
        
       
    try:
        wb.save(filename = path)
    except FileNotFoundError:
        print(f"The Folderpath {path} may not exists!!")
        sys.exit(1)


def read_body_radnet(body):
    newdict={}
    
    #Read till name of "Kontaktperson"
    newbody=body.replace('<br/>','')
    lines=newbody.splitlines()
    line_Kontaktperson=lines.index('Kontaktperson:')
    
    newdict['name_contact']=lines[line_Kontaktperson+2].lstrip()
    newdict['Verein_contact']=lines[line_Kontaktperson+3].lstrip()
    newdict['mail_contact']=lines[line_Kontaktperson+4].replace('Mail: ','').lstrip()
    

    line_Driver=lines.index('Angemeldete Fahrer:')
    newdict['race_class']=lines[line_Driver+3].replace('-','').strip()
    subline=lines[line_Driver+5].split(",")
    newdict['name_racer']=subline[0].lstrip()
    newdict['birth_racer']=subline[1].lstrip()
    newdict['Verein_racer']=subline[2].replace(' Verein/Team: ','').lstrip()
    newdict['licenc_racer']=subline[3].replace(' Lizenz: ','').lstrip()
    newdict['uciId_racer']=subline[4].replace(' UCI-Id: ','').lstrip()
    newdict['uciId_racer']=subline[5].replace(' UCI-Code: ','').lstrip()
    newdict['category_racer']=subline[6].replace(' Kategorie - Klasse: ','').lstrip()

    return newdict


def process_mails(M,oldlist,str_mail_header):
    """
    process all the mails in InBOx
    """
    newlist=[]
    rv, data = M.search(None, "ALL")
    if rv != 'OK':
        print("No messages found!")
        return

    for num in data[0].split():
        rv, data = M.fetch(num, '(RFC822)')
        if rv != 'OK':
            print("ERROR getting message", num)
            return

        msg = email.message_from_bytes(data[0][1])
        sendername=msg.get("From")
        recivedate =  msg.get("Date")
        emaiSubject = decode_header(msg.get("Subject"))[0][0]
        if isinstance(emaiSubject, bytes):
            # if it's a bytes, decode to str
            emaiSubject = emaiSubject.decode()
        


        if not str_mail_header == emaiSubject: #Check if Header are Equivilant
            continue

        if msg.is_multipart():
            for payload in msg.get_payload():
                print(payload)
        else:
            body=msg.get_payload()
            newlist.append(read_body_radnet(body))
    oldlist.extend(newlist)
    

def parse_yaml(opts):

    pth_config=''    
    for k,v in opts:
        if k == "-c":
            pth_config=v

    config=None
    with open(pth_config, 'r') as stream:
        try:
            config=yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)

    return config

def main(argv):

    if len(argv)<=1:
        print('Please write: main.py -c <config.yml>')
        sys.exit(1)

    try:
        opts, args = getopt.getopt(argv,"c:",["config"])
    except getopt.GetoptError:
        print('Please write: main.py -c <config.yml>')
        sys.exit(1)
    
    config=parse_yaml(opts)

    #login to Mail
    # create an IMAP4 class with SSL 
    str_imap_ssl=config['mail']['imap']
    print(f"Login to Server {str_imap_ssl}....")    
    try:        
        imap = imaplib.IMAP4_SSL(str_imap_ssl)
        print('DONE....')
    except:
        print("Error")
        sys.exit(1)
    
    
    # authenticate
    str_usr_name=config['mail']['usrname']
    str_usr_pw=config['mail']['passwort']    
    print(f"Login to Email: {str_usr_name}....")
    try:
        imap.login(str_usr_name, str_usr_pw)
        print('DONE....')
    except:
        print("Error")
        sys.exit(1)
    


    ls_mail_folders=config['mail']['folders']
    str_mail_header=config['mail']['standardheader']
    emaillist=[]
    for mail_folder in ls_mail_folders:
        print(f"processing Mailbox:{mail_folder}")
        rv, data = imap.select(mail_folder)
        if rv == 'OK':            
            process_mails(imap,emaillist,str_mail_header)
            print('DONE....')
        else:
            print("Error")
        
        

    imap.close()
    imap.logout()

    str_path=config['output']['path']
    str_filename=config['output']['name']
    if str_path is None:
        abspath=os.path.join(os.path.dirname(os.path.realpath(__file__)),str_filename)
    else:
        abspath=os.path.abspath(os.path.join(str_path,str_filename))
    print(f"Write Email Contents to File:{abspath}")
    try:
        parse_list_toxlsx(emaillist,abspath)
        print('DONE....')
    except:
        print("Error")
        
   

if __name__ == "__main__":
   main(sys.argv[1:])
